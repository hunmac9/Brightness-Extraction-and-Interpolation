import numpy as np
import logging
import re
import os
import csv
from datetime import datetime
from openpyxl import load_workbook

def get_sheet_data(excel_path):
    voltage_data = []
    current_data = []
    cycle_index_data = []

    try:
        workbook = load_workbook(excel_path, read_only=True, data_only=True)
        logging.info("Extracted Excel sheet information")

        for sheet_name in workbook.sheetnames:
            if re.match(r'^Channel_\d+_\d+$', sheet_name):
                sheet = workbook[sheet_name]
                
                rows = list(sheet.iter_rows(values_only=True))
                header = rows[0]

                try:
                    date_time_idx = header.index('Date_Time')
                    voltage_idx = header.index('Voltage(V)')
                    current_idx = header.index('Current(A)')
                    cycle_idx = header.index('Cycle_Index')
                    step_idx = header.index('Step_Index') if 'Step_Index' in header else None

                    for row in rows[1:]:
                        try:
                            date_time = row[date_time_idx]
                            if isinstance(date_time, str):
                                date_time = datetime.strptime(date_time, '%m/%d/%Y %H:%M:%S.%f')
                            voltage = row[voltage_idx]
                            current = row[current_idx]
                            cycle_index = row[cycle_idx]
                            if step_idx is not None and row[step_idx] == 1:
                                cycle_index = 0

                            voltage_data.append({'Date_Time': date_time, 'Voltage(V)': voltage})
                            current_data.append({'Date_Time': date_time, 'Current(A)': current})
                            cycle_index_data.append({'Date_Time': date_time, 'Cycle_Index': cycle_index})
                        except Exception as e:
                            logging.error(f"Failed to process row in {sheet_name}, check the data: {e}")
                except Exception as e:
                    logging.error(f"Failed to process {sheet_name}, check the data: {e}")
    except Exception as e:
        logging.error(f"Did not extract excel information: {e}")
        return None

    combined_data = {}
    for row in voltage_data:
        combined_data[row['Date_Time']] = {'Voltage(V)': row['Voltage(V)']}
    for row in current_data:
        if row['Date_Time'] in combined_data:
            combined_data[row['Date_Time']].update({'Current(A)': row['Current(A)']})
        else:
            combined_data[row['Date_Time']] = {'Current(A)': row['Current(A)']}
    for row in cycle_index_data:
        if row['Date_Time'] in combined_data:
            combined_data[row['Date_Time']].update({'Cycle_Index': row['Cycle_Index']})
        else:
            combined_data[row['Date_Time']] = {'Cycle_Index': row['Cycle_Index']}

    # Derive output directory names based on the directory where the input Excel file is located
    excel_dir = os.path.dirname(excel_path)
    output_file_name = os.path.join(excel_dir, "Echem_Extract.csv")
    
    try:
        with open(output_file_name, mode='w', newline='') as file:
            writer = csv.DictWriter(file, fieldnames=['Timestamp', 'Voltage(V)', 'Current(A)', 'Cycle_Index'])
            writer.writeheader()
            for timestamp, data in combined_data.items():
                row = {'Timestamp': timestamp}
                row.update(data)
                writer.writerow(row)
        
        logging.info(f"Data saved to {output_file_name}")
        return output_file_name
    except Exception as e:
        logging.error(f"Failed to save data to CSV: {e}")
        return None

if __name__ == "__main__":
    import sys
    logging.basicConfig(level=logging.INFO)
    
    if len(sys.argv) != 2:
        logging.error("Usage: python EchemProcessing.py <Excel_File_Path>")
        sys.exit(1)
    
    excel_path = sys.argv[1]
    
    if not os.path.exists(excel_path):
        logging.error(f"The file {excel_path} does not exist.")
        sys.exit(1)
    
    if not re.match(r'.*\.(xlsx|xlsm|xltx|xltm)$', excel_path, re.IGNORECASE):
        logging.error(f"The file {excel_path} does not have a supported extension.")
        sys.exit(1)
    
    try:
        workbook = load_workbook(excel_path, read_only=True, data_only=True)
        logging.info(f"Successfully opened the workbook: {excel_path}")
    except Exception as e:
        logging.error(f"Failed to open workbook: {e}")
        sys.exit(1)

    Echem_data_filepath = get_sheet_data(excel_path)
    if Echem_data_filepath:
        print(f"Echem data saved to: {Echem_data_filepath}")
    else:
        print("Failed to process and save data.")
